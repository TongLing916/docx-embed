from __future__ import annotations

from datetime import date, datetime, time
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from edp.common.markdown_normalizer import (
    normalize_markdown_cell_text,
    normalize_markdown_for_reading,
)
from edp.models import ParsedPackage, ParsedTable
from edp.xlsx.assets import XlsxAssetCollection, extract_xlsx_assets


def parse_xlsx_package(
    xlsx_path: str | Path, output_dir: str | Path, attachment_ref: str
) -> ParsedPackage:
    """Parse a single XLSX workbook into a structured attachment package.

    Features:

    * Dual-workbook loading: formulas in one workbook (``data_only=False``),
      computed values in another (``data_only=True``).
    * Sub-table splitting: repeated header rows, blank-row gaps, and image
      anchors are used as boundary signals.
    * Cell metadata: formulas, hyperlinks, comments, number formats, data types.
    * Merged-cell ranges and hidden row/column annotations.
    * Asset extraction: images, charts, OLE objects, and OMML equations are
      extracted with anchor metadata.
    * Framework enrichment: when ``unstructured`` is installed, its xlsx
      partitioner runs as a sidecar for comparison.
    * Chinese-aware Markdown preview that adapts labels to the content language.
    """

    source = Path(xlsx_path)
    package_dir = Path(output_dir)
    tables_dir = package_dir / "tables"
    tables_dir.mkdir(parents=True, exist_ok=True)

    warnings: list[str] = []
    tables: list[ParsedTable] = []
    workbook = load_workbook(source, data_only=False)
    values_workbook = load_workbook(source, data_only=True)

    assets_collection = extract_xlsx_assets(source, package_dir / "assets")
    warnings.extend(assets_collection.warnings)

    for sheet in workbook.worksheets:
        value_sheet = values_workbook[sheet.title]
        sheet_assets = _assets_for_sheet(assets_collection, sheet.title)
        sub_tables = _split_sheet_into_tables(sheet, value_sheet, sheet_assets)
        if not sub_tables:
            continue

        for sub in sub_tables:
            # Widen end_row to include assets anchored beyond data rows.
            max_asset_row = _max_asset_row(sheet_assets)
            if max_asset_row > sub["end_row"]:
                sub["end_row"] = max_asset_row

            table_index = len(tables) + 1
            csv_path = tables_dir / f"table_{table_index:03d}.csv"
            json_path = tables_dir / f"table_{table_index:03d}.json"

            with csv_path.open("w", newline="", encoding="utf-8") as stream:
                writer = csv.writer(stream)
                for row in sub["rows"]:
                    writer.writerow(
                        ["" if value is None else value for value in row]
                    )

            table_json = {
                "sheet": sheet.title,
                "state": sheet.sheet_state,
                "start_row": sub["start_row"],
                "end_row": sub["end_row"],
                "boundary_signals": sub["boundary_signals"],
                "rows": sub["rows"],
                "cells": _filter_cells_by_row(
                    _sheet_cell_metadata(sheet, value_sheet),
                    sub["start_row"],
                    sub["end_row"],
                ),
                "merged_ranges": _filter_merged_ranges_by_row(
                    _merged_ranges(sheet), sub["start_row"], sub["end_row"]
                ),
                "hidden_rows": _filter_hidden_rows(
                    _hidden_rows(sheet), sub["start_row"], sub["end_row"]
                ),
                "hidden_columns": _hidden_columns(sheet),
                "chart_count": sum(
                    1
                    for c in sheet_assets.get("charts", [])
                    if _asset_in_row_range(c, sub["start_row"], sub["end_row"])
                ),
                "image_count": sum(
                    1
                    for img in sheet_assets.get("images", [])
                    if _asset_in_row_range(img, sub["start_row"], sub["end_row"])
                ),
                "assets": _filter_assets_by_row(
                    sheet_assets, sub["start_row"], sub["end_row"]
                ),
                "framework_result": _framework_result_placeholder(),
            }
            json_path.write_text(
                json.dumps(table_json, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            tables.append(
                ParsedTable(
                    sheet_name=sheet.title,
                    csv_path=csv_path,
                    json_path=json_path,
                    row_count=len(sub["rows"]),
                )
            )

    if not tables:
        warnings.append(f"Workbook {source.name} did not contain non-empty sheets")

    metadata_path = package_dir / "workbook.json"
    metadata_path.write_text(
        json.dumps(
            _workbook_metadata(source, workbook, assets_collection),
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    content_path = package_dir / "content.md"
    content_path.write_text(
        normalize_markdown_for_reading(
            _render_workbook_markdown(
                attachment_ref, source.name, tables, metadata_path
            )
        ),
        encoding="utf-8",
    )

    _enrich_with_frameworks(source, tables, package_dir, warnings)

    return ParsedPackage(
        ref=attachment_ref,
        package_dir=package_dir,
        content_path=content_path,
        tables=tables,
        warnings=warnings,
    )


# --------------------------------------------------------------------------- #
# Sheet splitting
# --------------------------------------------------------------------------- #


def _split_sheet_into_tables(
    sheet, value_sheet, sheet_assets: dict[str, list]
) -> list[dict]:
    """Split a worksheet into sub-tables using image anchors, repeated
    header-row signatures, and blank-row gaps as boundary signals.

    Returns a list of dicts, each with ``rows``, ``start_row``,
    ``end_row``, and ``boundary_signals``. Row numbers are 1-indexed.
    """

    indexed = _indexed_rows(value_sheet)
    image_rows = _image_anchor_rows(sheet_assets)
    header_rows = _repeated_header_rows(indexed)
    blank_rows = _blank_gap_boundaries(indexed)

    # Image anchors only confirm boundaries detected by other methods;
    # standalone images inside a continuous table must not force a split.
    structural_boundaries = header_rows | blank_rows
    all_boundaries = sorted(structural_boundaries)

    sub_tables: list[dict] = []
    start = 1
    for boundary in all_boundaries:
        if boundary <= start:
            continue
        table_rows = [
            (row_idx, values)
            for row_idx, values in indexed
            if start <= row_idx < boundary
            and values
            and any(v not in {None, ""} for v in values)
        ]
        if table_rows:
            signals: list[str] = []
            if boundary in image_rows:
                signals.append("image_anchor")
            if boundary in header_rows:
                signals.append("header_repeat")
            if boundary in blank_rows:
                signals.append("blank_row_gap")
            signals.sort()
            sub_tables.append(
                {
                    "rows": _prune_blank_rows_and_columns(
                        [values for _, values in table_rows]
                    ),
                    "start_row": min(r for r, _ in table_rows),
                    "end_row": max(r for r, _ in table_rows),
                    "boundary_signals": signals if signals else [],
                }
            )
        start = boundary

    # Final table after the last boundary.
    table_rows = [
        (row_idx, values)
        for row_idx, values in indexed
        if row_idx >= start
        and values
        and any(v not in {None, ""} for v in values)
    ]
    if table_rows:
        sub_tables.append(
            {
                "rows": _prune_blank_rows_and_columns(
                    [values for _, values in table_rows]
                ),
                "start_row": min(r for r, _ in table_rows),
                "end_row": max(r for r, _ in table_rows),
                "boundary_signals": [],
            }
        )

    return sub_tables


def _indexed_rows(sheet) -> list[tuple[int, list[object | None]]]:
    """Return (row_number, values) for every row, keeping empty rows."""

    indexed: list[tuple[int, list[object | None]]] = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_json_safe_value(value) for value in row]
        while values and values[-1] is None:
            values.pop()
        indexed.append((row_idx, values))
    return indexed


def _max_asset_row(sheet_assets: dict[str, list]) -> int:
    """Return the highest 1-indexed row where any asset is anchored."""

    max_row = 0
    for asset_list in sheet_assets.values():
        for asset in asset_list:
            anchor_row = asset.get("from_row")
            if isinstance(anchor_row, (int, float)) and anchor_row >= 0:
                max_row = max(max_row, int(anchor_row) + 1)
    return max_row


def _image_anchor_rows(sheet_assets: dict[str, list]) -> set[int]:
    """Collect 1-indexed rows where images or charts are anchored."""

    rows: set[int] = set()
    for asset_list in sheet_assets.values():
        for asset in asset_list:
            anchor_row = asset.get("from_row")
            if isinstance(anchor_row, (int, float)) and anchor_row >= 0:
                rows.add(int(anchor_row) + 1)
    return rows


def _repeated_header_rows(
    indexed: list[tuple[int, list[object | None]]],
) -> set[int]:
    """Find rows whose non-empty value signature appears ≥3 times across
    the sheet AND whose content is predominantly text (not numbers).

    These are likely recurring table headers.
    """

    sig_counts: dict[tuple[str, ...], int] = {}
    for _row_idx, values in indexed:
        if not values:
            continue
        signature = tuple(
            str(v) for v in values if v is not None and v != ""
        )
        if not signature:
            continue
        sig_counts[signature] = sig_counts.get(signature, 0) + 1

    header_sigs: set[tuple[str, ...]] = set()
    for sig, count in sig_counts.items():
        if count < 3:
            continue
        text_cells = sum(
            1
            for v in sig
            if isinstance(v, str) and not _looks_numeric(v)
        )
        if text_cells >= 2:
            header_sigs.add(sig)

    boundary_rows: set[int] = set()
    for row_idx, values in indexed:
        if not values:
            continue
        signature = tuple(
            str(v) for v in values if v is not None and v != ""
        )
        if signature in header_sigs:
            boundary_rows.add(row_idx)

    return boundary_rows


def _looks_numeric(value: str) -> bool:
    """Return True when *value* looks like a plain number."""

    stripped = value.strip()
    if not stripped:
        return False
    try:
        float(stripped.replace("%", "").replace(",", ""))
        return True
    except ValueError:
        return False


def _blank_gap_boundaries(
    indexed: list[tuple[int, list[object | None]]],
) -> set[int]:
    """Return rows immediately after ≥2 consecutive fully-empty rows."""

    boundaries: set[int] = set()
    blank_run = 0
    for row_idx, values in indexed:
        if not values or all(v in {None, ""} for v in values):
            blank_run += 1
        else:
            if blank_run >= 2:
                boundaries.add(row_idx)
            blank_run = 0
    return boundaries


def _prune_blank_rows_and_columns(
    rows: list[list[object | None]],
) -> list[list[object | None]]:
    """Remove entirely blank rows and columns from a sub-table."""

    kept = [
        row for row in rows if any(v not in {None, ""} for v in row)
    ]
    if not kept:
        return kept

    max_width = max(len(row) for row in kept)
    blank_columns: set[int] = set()
    for col_idx in range(max_width):
        if all(
            col_idx >= len(row) or row[col_idx] in {None, ""}
            for row in kept
        ):
            blank_columns.add(col_idx)

    if not blank_columns:
        return kept

    pruned: list[list[object | None]] = []
    for row in kept:
        pruned_row = [
            row[col_idx] if col_idx < len(row) else None
            for col_idx in range(max_width)
            if col_idx not in blank_columns
        ]
        while pruned_row and pruned_row[-1] is None:
            pruned_row.pop()
        pruned.append(pruned_row)
    return pruned


# --------------------------------------------------------------------------- #
# Cell metadata
# --------------------------------------------------------------------------- #


def _sheet_cell_metadata(sheet, value_sheet) -> dict[str, dict[str, object]]:
    cells: dict[str, dict[str, object]] = {}
    for row in sheet.iter_rows():
        for cell in row:
            value_cell = value_sheet[cell.coordinate]
            value = _json_safe_value(value_cell.value)
            formula = (
                cell.value
                if isinstance(cell.value, str) and cell.value.startswith("=")
                else None
            )
            has_content = any(
                item not in {None, ""}
                for item in (
                    value,
                    formula,
                    cell.hyperlink.target if cell.hyperlink else None,
                    cell.comment.text if cell.comment else None,
                )
            )
            if not has_content:
                continue
            entry: dict[str, object] = {
                "value": value,
                "data_type": cell.data_type,
                "number_format": cell.number_format,
            }
            if formula is not None:
                entry["formula"] = formula
                entry["display_value"] = value
            if cell.hyperlink:
                entry["hyperlink"] = cell.hyperlink.target
            if cell.comment:
                entry["comment"] = cell.comment.text
            cells[cell.coordinate] = entry
    return cells


def _merged_ranges(sheet) -> list[str]:
    return [str(cell_range) for cell_range in sheet.merged_cells.ranges]


def _hidden_rows(sheet) -> list[int]:
    return [
        index
        for index, dimension in sorted(sheet.row_dimensions.items())
        if getattr(dimension, "hidden", False)
    ]


def _hidden_columns(sheet) -> list[str]:
    hidden: list[str] = []
    for key, dimension in sorted(sheet.column_dimensions.items()):
        if not getattr(dimension, "hidden", False):
            continue
        if isinstance(key, int):
            hidden.append(get_column_letter(key))
        else:
            hidden.append(str(key))
    return hidden


# --------------------------------------------------------------------------- #
# Cell/row filtering helpers
# --------------------------------------------------------------------------- #


def _cell_row_number(coordinate: str) -> int:
    """Extract the 1-indexed row from a coordinate like ``'A2'`` or ``'AB123'``."""

    match = re.match(r"^[A-Z]+(\d+)$", coordinate)
    if match:
        return int(match.group(1))
    return 0


def _filter_cells_by_row(
    cells: dict[str, dict[str, object]],
    start_row: int,
    end_row: int,
) -> dict[str, dict[str, object]]:
    """Return only cells whose row number falls in ``[start_row, end_row]``."""

    return {
        coord: entry
        for coord, entry in cells.items()
        if start_row <= _cell_row_number(coord) <= end_row
    }


def _filter_merged_ranges_by_row(
    ranges: list[str],
    start_row: int,
    end_row: int,
) -> list[str]:
    kept: list[str] = []
    for rng in ranges:
        try:
            start, end = rng.split(":")
            r_start = _cell_row_number(start)
            r_end = _cell_row_number(end)
        except ValueError:
            continue
        if r_start <= end_row and r_end >= start_row:
            kept.append(rng)
    return kept


def _filter_hidden_rows(
    hidden: list[int],
    start_row: int,
    end_row: int,
) -> list[int]:
    return [r for r in hidden if start_row <= r <= end_row]


def _asset_in_row_range(
    asset: dict, start_row: int, end_row: int
) -> bool:
    """Return True when *asset*'s anchor row is within ``[start_row, end_row]``."""

    anchor_row = asset.get("from_row")
    if not isinstance(anchor_row, (int, float)):
        return False
    row_1indexed = int(anchor_row) + 1
    return start_row <= row_1indexed <= end_row


def _filter_assets_by_row(
    sheet_assets: dict[str, list],
    start_row: int,
    end_row: int,
) -> dict[str, list]:
    """Filter asset records to those anchored in ``[start_row, end_row]``."""

    return {
        key: [
            record
            for record in records
            if _asset_in_row_range(record, start_row, end_row)
        ]
        for key, records in sheet_assets.items()
    }


def _assets_for_sheet(
    assets: XlsxAssetCollection, sheet_name: str
) -> dict[str, list]:
    """Filter the asset collection down to records anchored on *sheet_name*."""

    def belonging(records: list[dict]) -> list[dict]:
        return [
            record
            for record in records
            if record.get("sheet") == sheet_name
        ]

    return {
        "images": belonging(assets.images),
        "charts": belonging(assets.charts),
        "attachments": belonging(assets.attachments),
        "equations": belonging(assets.equations),
    }


# --------------------------------------------------------------------------- #
# Framework enrichment (optional)
# --------------------------------------------------------------------------- #


def _framework_result_placeholder() -> dict[str, object | None]:
    return {"unstructured": None, "docling": None, "mineru": None}


def _enrich_with_frameworks(
    source: Path,
    tables: list[ParsedTable],
    package_dir: Path,
    warnings: list[str],
) -> None:
    """Run available framework parsers on *source* and store results as
    sidecar data in each table JSON and in ``workbook.json``.

    Framework failures are silent — no exceptions, no warnings.
    """

    framework_results: dict[str, object | None] = {}

    unstructured_result = _run_unstructured(source)
    if unstructured_result is not None:
        framework_results["unstructured"] = unstructured_result

    if not framework_results:
        return

    for table in tables:
        try:
            table_json = json.loads(
                table.json_path.read_text(encoding="utf-8")
            )
        except (OSError, json.JSONDecodeError):
            continue

        table_framework: dict[str, object | None] = {}
        if unstructured_result is not None:
            matched = _match_unstructured_elements(
                unstructured_result,
                table_json.get("start_row"),
                table_json.get("end_row"),
            )
            table_framework["unstructured"] = matched
        table_framework["docling"] = None
        table_framework["mineru"] = None
        table_json["framework_result"] = table_framework
        try:
            table.json_path.write_text(
                json.dumps(table_json, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except OSError:
            pass

    metadata_path = package_dir / "workbook.json"
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return
    metadata["framework_results"] = framework_results
    try:
        metadata_path.write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        pass


def _run_unstructured(source: Path) -> list[dict[str, object]] | None:
    """Run ``unstructured.partition.xlsx.partition_xlsx`` and return a
    list of element dicts, or ``None`` on any failure."""

    try:
        from unstructured.documents.elements import Table
        from unstructured.partition.xlsx import partition_xlsx
    except ImportError:
        return None

    try:
        elements = partition_xlsx(filename=str(source))
    except Exception:
        return None

    results: list[dict[str, object]] = []
    for element in elements:
        record: dict[str, object] = {
            "type": type(element).__name__,
            "text": (
                str(element.metadata.text_as_html)
                if isinstance(element, Table)
                else str(element)
            ),
        }
        meta = getattr(element, "metadata", None)
        if meta is not None:
            record["metadata"] = {
                k: str(v) for k, v in meta.to_dict().items()
            }
        results.append(record)
    return results


def _match_unstructured_elements(
    unstructured_result: list[dict[str, object]],
    start_row: object,
    end_row: object,
) -> dict[str, object]:
    """Return a summary of unstructured elements that overlap our sub-table.

    When row information is unavailable, report ``matched_count`` as the
    total number of detected elements.
    """

    element_count = len(unstructured_result)
    table_elements = [
        e for e in unstructured_result if e.get("type") == "Table"
    ]
    return {
        "element_count": element_count,
        "table_count": len(table_elements),
        "matched_count": element_count,
        "note": (
            "element-to-row matching is best-effort "
            "without per-element row metadata"
        ),
    }


# --------------------------------------------------------------------------- #
# Workbook metadata
# --------------------------------------------------------------------------- #


def _workbook_metadata(
    source: Path, workbook, assets: XlsxAssetCollection
) -> dict[str, object]:
    return {
        "source_file": source.name,
        "sheet_count": len(workbook.worksheets),
        "sheets": [
            _sheet_metadata(sheet, _assets_for_sheet(assets, sheet.title))
            for sheet in workbook.worksheets
        ],
        "assets": {
            "images": assets.images,
            "charts": assets.charts,
            "attachments": assets.attachments,
            "equations": assets.equations,
        },
    }


def _sheet_metadata(
    sheet, sheet_assets: dict[str, list]
) -> dict[str, object]:
    return {
        "name": sheet.title,
        "state": sheet.sheet_state,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_ranges": _merged_ranges(sheet),
        "hidden_rows": _hidden_rows(sheet),
        "hidden_columns": _hidden_columns(sheet),
        "chart_count": len(getattr(sheet, "_charts", [])),
        "image_count": len(getattr(sheet, "_images", [])),
        "assets": sheet_assets,
    }


# --------------------------------------------------------------------------- #
# Markdown rendering
# --------------------------------------------------------------------------- #


def _is_predominantly_chinese(rows: list[list[object | None]]) -> bool:
    """Return True when the majority of text cells contain Chinese characters."""

    text_chars = 0
    chinese_chars = 0
    for row in rows:
        for cell in row:
            if not isinstance(cell, str) or not cell.strip():
                continue
            for ch in cell:
                text_chars += 1
                if "\u4e00" <= ch <= "\u9fff":
                    chinese_chars += 1
    return text_chars > 0 and chinese_chars / text_chars >= 0.3


def _render_workbook_markdown(
    attachment_ref: str,
    source_name: str,
    tables: list[ParsedTable],
    metadata_path: Path | None = None,
) -> str:
    lines = [
        f"# Embedded Workbook {attachment_ref}",
        "",
        f"Source file: `{source_name}`",
        "",
    ]
    if metadata_path is not None:
        lines.extend([f"- Metadata: `{metadata_path.name}`", ""])

    if not tables:
        lines.append("No non-empty sheets were found.")
        lines.append("")
        return "\n".join(lines)

    # Group tables by sheet to number them (e.g. "Table 2/5").
    sheet_table_counts: dict[str, int] = {}
    sheet_table_indices: dict[str, int] = {}
    for tbl in tables:
        sheet_table_counts[tbl.sheet_name] = (
            sheet_table_counts.get(tbl.sheet_name, 0) + 1
        )

    for table in tables:
        sheet_name = table.sheet_name
        sheet_table_indices[sheet_name] = (
            sheet_table_indices.get(sheet_name, 0) + 1
        )
        idx = sheet_table_indices[sheet_name]
        total = sheet_table_counts[sheet_name]
        heading = f"## Sheet: {sheet_name}"
        if total > 1:
            heading += f" · Table {idx}/{total}"
        lines.extend(
            [
                heading,
                "",
                f"- CSV: `tables/{table.csv_path.name}`",
                f"- JSON: `tables/{table.json_path.name}`",
                "",
            ]
        )
        table_json = json.loads(table.json_path.read_text(encoding="utf-8"))
        rows = table_json["rows"]
        use_chinese = _is_predominantly_chinese(rows)
        lines.extend(_markdown_preview(rows))
        assets = table_json.get("assets") or {}
        asset_lines = _assets_markdown(assets, use_chinese=use_chinese)
        if asset_lines:
            lines.append("")
            lines.extend(asset_lines)
        lines.append("")

    return "\n".join(lines)


def _assets_markdown(
    assets: dict[str, list], *, use_chinese: bool = False
) -> list[str]:
    """Render per-sheet images, charts, equations and embedded objects."""

    lines: list[str] = []
    _at = " 位于 " if use_chinese else " at "
    _caption_label = "标题: " if use_chinese else "Caption: "
    _ocr_label = "OCR文字: " if use_chinese else "OCR: "
    _desc_label = "描述文件: " if use_chinese else "Description: "
    _chart_see = "详见 " if use_chinese else "see "
    _equation_label = "公式" if use_chinese else "Equation"
    _embedded_label = "嵌入对象" if use_chinese else "Embedded"
    _untitled = "(无标题)" if use_chinese else "(untitled)"

    for image in assets.get("images", []):
        where = f"{_at}`{image['cell']}`" if image.get("cell") else ""
        lines.append(f"![{image['ref']}](assets/{image['path']}){where}")
        caption = image.get("caption")
        if caption:
            lines.append(f"  - {_caption_label}{caption}")
        ocr_text = image.get("ocr_text")
        if ocr_text:
            lines.append(f"  - {_ocr_label}{ocr_text}")
        description_path = image.get("description_path")
        if description_path:
            lines.append(f"  - {_desc_label}`assets/{description_path}`")
    for chart in assets.get("charts", []):
        title = chart.get("title") or _untitled
        where = f"{_at}`{chart['cell']}`" if chart.get("cell") else ""
        lines.append(
            f"- Chart `{chart['ref']}`{where}: {title} "
            f"({_chart_see}`assets/{chart['json_path']}`)"
        )
    for equation in assets.get("equations", []):
        where = f"{_at}`{equation['cell']}`" if equation.get("cell") else ""
        lines.append(
            f"- {_equation_label}{where}: `{equation['text']}`"
        )
    for attachment in assets.get("attachments", []):
        where = (
            f"{_at}`{attachment['cell']}`" if attachment.get("cell") else ""
        )
        lines.append(
            f"- {_embedded_label} `{attachment['ref']}` "
            f"({attachment['type']}){where} "
            f"({_chart_see}`assets/{attachment['path']}`)"
        )
    return lines


def _markdown_preview(rows: list[list[object | None]]) -> list[str]:
    """Render *all* rows as a GFM table; no truncation."""

    if not rows:
        return []

    max_width = max(len(row) for row in rows)
    padded = [row + [None] * (max_width - len(row)) for row in rows]
    rendered = [
        "| " + " | ".join(_markdown_cell(value) for value in padded[0]) + " |"
    ]
    rendered.append(
        "| " + " | ".join("---" for _ in range(max_width)) + " |"
    )
    for row in padded[1:]:
        rendered.append(
            "| "
            + " | ".join(_markdown_cell(value) for value in row)
            + " |"
        )
    return rendered


def _markdown_cell(value: object | None) -> str:
    if value is None:
        return ""
    # Markdown table cells cannot contain literal newlines — a ``\n`` inside a
    # value would break the row across multiple physical lines and shatter the
    # table. Escape ``|`` (column separator) and render embedded newlines as
    # ``<br>``, the standard way to keep intra-cell line breaks in a GFM table
    # without losing them. Raw rows/cells/CSV/JSON keep the original ``\n``.
    return normalize_markdown_cell_text(str(value)).replace("|", "\\|")


# --------------------------------------------------------------------------- #
# Shared utility
# --------------------------------------------------------------------------- #


def _json_safe_value(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value
