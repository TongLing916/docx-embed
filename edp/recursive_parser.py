from __future__ import annotations

from datetime import date, datetime, time
import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from edp.models import EmbeddedObject, ParsedPackage, ParsedTable


def parse_attachment_package(embedded: EmbeddedObject, output_dir: str | Path) -> ParsedPackage:
    """Parse supported embedded attachments into structured child packages."""

    if embedded.type == "xlsx":
        return parse_xlsx_package(embedded.path, output_dir, embedded.ref)
    raise ValueError(f"Unsupported attachment type for parsing: {embedded.type}")


def parse_xlsx_package(
    xlsx_path: str | Path, output_dir: str | Path, attachment_ref: str
) -> ParsedPackage:
    """Parse a single XLSX workbook into a small structured attachment package."""

    source = Path(xlsx_path)
    package_dir = Path(output_dir)
    tables_dir = package_dir / "tables"
    tables_dir.mkdir(parents=True, exist_ok=True)

    warnings: list[str] = []
    tables: list[ParsedTable] = []
    workbook = load_workbook(source, data_only=True)

    for sheet in workbook.worksheets:
        rows = _non_empty_rows(sheet)
        if not rows:
            continue

        table_index = len(tables) + 1
        csv_path = tables_dir / f"table_{table_index:03d}.csv"
        json_path = tables_dir / f"table_{table_index:03d}.json"

        with csv_path.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.writer(stream)
            for row in rows:
                writer.writerow(["" if value is None else value for value in row])

        json_path.write_text(
            json.dumps({"sheet": sheet.title, "rows": rows}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        tables.append(
            ParsedTable(
                sheet_name=sheet.title,
                csv_path=csv_path,
                json_path=json_path,
                row_count=len(rows),
            )
        )

    if not tables:
        warnings.append(f"Workbook {source.name} did not contain non-empty sheets")

    content_path = package_dir / "content.md"
    content_path.write_text(_render_workbook_markdown(attachment_ref, source.name, tables), encoding="utf-8")

    return ParsedPackage(
        ref=attachment_ref,
        package_dir=package_dir,
        content_path=content_path,
        tables=tables,
        warnings=warnings,
    )


def _non_empty_rows(sheet) -> list[list[object | None]]:
    rows: list[list[object | None]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [_json_safe_value(value) for value in row]
        while values and values[-1] is None:
            values.pop()
        if any(value not in {None, ""} for value in values):
            rows.append(values)
    return rows


def _json_safe_value(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value


def _render_workbook_markdown(
    attachment_ref: str, source_name: str, tables: list[ParsedTable]
) -> str:
    lines = [
        f"# Embedded Workbook {attachment_ref}",
        "",
        f"Source file: `{source_name}`",
        "",
    ]

    if not tables:
        lines.append("No non-empty sheets were found.")
        lines.append("")
        return "\n".join(lines)

    for table in tables:
        lines.extend(
            [
                f"## Sheet: {table.sheet_name}",
                "",
                f"- CSV: `tables/{table.csv_path.name}`",
                f"- JSON: `tables/{table.json_path.name}`",
                "",
            ]
        )
        rows = json.loads(table.json_path.read_text(encoding="utf-8"))["rows"]
        lines.extend(_markdown_preview(rows))
        lines.append("")

    return "\n".join(lines)


def _markdown_preview(rows: list[list[object | None]]) -> list[str]:
    preview_rows = rows[:10]
    if not preview_rows:
        return []

    max_width = max(len(row) for row in preview_rows)
    padded = [row + [None] * (max_width - len(row)) for row in preview_rows]
    rendered = ["| " + " | ".join(_markdown_cell(value) for value in padded[0]) + " |"]
    rendered.append("| " + " | ".join("---" for _ in range(max_width)) + " |")
    for row in padded[1:]:
        rendered.append("| " + " | ".join(_markdown_cell(value) for value in row) + " |")
    return rendered


def _markdown_cell(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|")
