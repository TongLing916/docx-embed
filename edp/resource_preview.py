from __future__ import annotations

from dataclasses import replace
from datetime import date, datetime, time
import json
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from edp.models import EmbeddedObject, ExtractionResult


TEXT_PREVIEW_BYTES = 100 * 1024
XLSX_PREVIEW_ROWS = 10
TEXT_PREVIEW_TYPES = {"txt", "log", "csv"}
CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
CHART_XML_NS = {"c": CHART_NS, "a": DRAWING_NS}
DIAGRAM_NS = "http://schemas.openxmlformats.org/drawingml/2006/diagram"
DIAGRAM_XML_NS = {"dgm": DIAGRAM_NS, "a": DRAWING_NS}


def build_resource_previews(
    extraction: ExtractionResult, output_dir: str | Path
) -> ExtractionResult:
    """Create previews and return extraction metadata updated with parse status."""

    package_dir = Path(output_dir)
    resources_dir = package_dir / "structured" / "resources"
    warnings = [*extraction.warnings]

    objects: list[EmbeddedObject] = []
    for embedded in extraction.objects:
        updated, preview_warnings = _attachment_with_preview(embedded, resources_dir)
        objects.append(updated)
        warnings.extend(preview_warnings)

    images = [_image_with_preserved_status(image) for image in extraction.images]
    charts: list[EmbeddedObject] = []
    for chart in extraction.charts:
        updated, preview_warnings = _chart_with_preview(chart, resources_dir)
        charts.append(updated)
        warnings.extend(preview_warnings)
    diagrams: list[EmbeddedObject] = []
    for diagram in extraction.diagrams:
        updated, preview_warnings = _diagram_with_preview(diagram, resources_dir)
        diagrams.append(updated)
        warnings.extend(preview_warnings)
    return replace(
        extraction,
        objects=objects,
        images=images,
        charts=charts,
        diagrams=diagrams,
        warnings=warnings,
    )


def _attachment_with_preview(
    embedded: EmbeddedObject, resources_dir: Path
) -> tuple[EmbeddedObject, list[str]]:
    attachment_type = embedded.type.lower()
    if attachment_type in TEXT_PREVIEW_TYPES:
        return _text_attachment_with_preview(embedded, resources_dir)
    if attachment_type == "xlsx":
        return _xlsx_attachment_with_preview(embedded, resources_dir)
    return _attachment_extracted_only(embedded), []


def _text_attachment_with_preview(
    embedded: EmbeddedObject, resources_dir: Path
) -> tuple[EmbeddedObject, list[str]]:
    preview_path = resources_dir / embedded.ref / "preview.md"
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    payload = embedded.path.read_bytes()
    preview_payload = payload[:TEXT_PREVIEW_BYTES]
    text = preview_payload.decode("utf-8", errors="replace")
    if len(payload) > TEXT_PREVIEW_BYTES:
        text += "\n\n[Preview truncated at 100KB]"
    preview_path.write_text(
        f"# Resource Preview {embedded.ref}\n\n```text\n{text}\n```\n",
        encoding="utf-8",
    )
    return (
        replace(
            embedded,
            parse_policy=_preview_policy(),
            parse_status=_preview_status(embedded.ref, json_path=None),
        ),
        [],
    )


def _xlsx_attachment_with_preview(
    embedded: EmbeddedObject, resources_dir: Path
) -> tuple[EmbeddedObject, list[str]]:
    preview_dir = resources_dir / embedded.ref
    preview_md = preview_dir / "preview.md"
    preview_json = preview_dir / "preview.json"
    preview_dir.mkdir(parents=True, exist_ok=True)

    try:
        workbook = load_workbook(
            embedded.path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        return (
            _attachment_extracted_only(embedded),
            [f"Failed to preview {embedded.filename}: {exc}"],
        )

    try:
        sheets = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(max_row=XLSX_PREVIEW_ROWS, values_only=True):
                rows.append(_trim_row([_json_safe_value(value) for value in row]))
            sheets.append(
                {
                    "name": sheet.title,
                    "max_row": sheet.max_row or 0,
                    "max_column": sheet.max_column or 0,
                    "preview_rows": rows,
                }
            )
    finally:
        workbook.close()

    preview_data = {
        "ref": embedded.ref,
        "filename": embedded.filename,
        "type": "xlsx",
        "sheets": sheets,
    }
    preview_json.write_text(
        json.dumps(preview_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    preview_md.write_text(
        _render_xlsx_preview(embedded, sheets),
        encoding="utf-8",
    )
    return (
        replace(
            embedded,
            parse_policy=_preview_policy(),
            parse_status=_preview_status(embedded.ref, json_path=preview_json),
        ),
        [],
    )


def _attachment_extracted_only(embedded: EmbeddedObject) -> EmbeddedObject:
    return replace(
        embedded,
        parse_policy={
            "mode": "extracted_only",
            "reason": "preview disabled for this attachment type",
        },
        parse_status={
            "status": "extracted_only",
            "preview_path": None,
            "preview_json_path": None,
            "full_parse": False,
        },
    )


def _image_with_preserved_status(image: EmbeddedObject) -> EmbeddedObject:
    return replace(
        image,
        parse_policy={
            "mode": "preserved_image",
            "reason": "image asset preserved without OCR",
        },
        parse_status={
            "status": "preserved",
            "preview_path": None,
            "preview_json_path": None,
            "full_parse": False,
        },
    )


def _chart_with_preview(chart: EmbeddedObject, resources_dir: Path) -> tuple[EmbeddedObject, list[str]]:
    preview_dir = resources_dir / chart.ref
    preview_md = preview_dir / "preview.md"
    preview_json = preview_dir / "preview.json"
    preview_dir.mkdir(parents=True, exist_ok=True)

    try:
        chart_data = _read_chart_data(chart)
    except Exception as exc:
        return (
            replace(
                chart,
                parse_policy={
                    "mode": "extracted_only",
                    "reason": "chart preview extraction failed",
                },
                parse_status={
                    "status": "extracted_only",
                    "preview_path": None,
                    "preview_json_path": None,
                    "full_parse": False,
                },
            ),
            [f"Failed to preview {chart.filename}: {exc}"],
        )

    preview_json.write_text(json.dumps(chart_data, ensure_ascii=False, indent=2), encoding="utf-8")
    preview_md.write_text(_render_chart_preview(chart, chart_data), encoding="utf-8")
    return (
        replace(
            chart,
            parse_policy={
                "mode": "shallow_preview",
                "reason": "allowlisted chart preview",
            },
            parse_status=_preview_status(chart.ref, json_path=preview_json),
        ),
        [],
    )


def _preview_policy() -> dict[str, object]:
    return {
        "mode": "shallow_preview",
        "reason": "allowlisted attachment preview",
    }


def _diagram_with_preview(
    diagram: EmbeddedObject, resources_dir: Path
) -> tuple[EmbeddedObject, list[str]]:
    preview_dir = resources_dir / diagram.ref
    preview_md = preview_dir / "preview.md"
    preview_json = preview_dir / "preview.json"
    preview_dir.mkdir(parents=True, exist_ok=True)

    try:
        diagram_data = _read_diagram_data(diagram)
    except Exception as exc:
        return (
            replace(
                diagram,
                parse_policy={
                    "mode": "extracted_only",
                    "reason": "SmartArt preview extraction failed",
                },
                parse_status={
                    "status": "extracted_only",
                    "preview_path": None,
                    "preview_json_path": None,
                    "full_parse": False,
                },
            ),
            [f"Failed to preview {diagram.filename}: {exc}"],
        )

    preview_json.write_text(
        json.dumps(diagram_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    preview_md.write_text(_render_diagram_preview(diagram, diagram_data), encoding="utf-8")
    return (
        replace(
            diagram,
            parse_policy={
                "mode": "shallow_preview",
                "reason": "allowlisted SmartArt preview",
            },
            parse_status=_preview_status(diagram.ref, json_path=preview_json),
        ),
        [],
    )


def _read_diagram_data(diagram: EmbeddedObject) -> dict[str, object]:
    root = ET.fromstring(diagram.path.read_bytes())
    texts = _unique_preserving_order(_texts(root, ".//a:t", DIAGRAM_XML_NS))
    return {
        "ref": diagram.ref,
        "filename": diagram.filename,
        "type": "diagram",
        "source_path": diagram.source_path,
        "texts": texts,
        "companion_parts": diagram.related_parts,
    }


def _read_chart_data(chart: EmbeddedObject) -> dict[str, object]:
    root = ET.fromstring(chart.path.read_bytes())
    series_data = []
    categories: list[str] = []
    for ser in root.findall(".//c:ser", CHART_XML_NS):
        name = _first_text(ser, ".//c:tx//c:v") or _first_text(ser, ".//c:tx//a:t")
        ser_categories = _texts(ser, ".//c:cat//c:v")
        values = [_number_or_text(text) for text in _texts(ser, ".//c:val//c:v")]
        if ser_categories and not categories:
            categories = ser_categories
        series_data.append({"name": name or f"Series {len(series_data) + 1}", "values": values})

    rows = []
    for index, category in enumerate(categories):
        values = {}
        for series in series_data:
            series_values = series["values"]
            values[str(series["name"])] = series_values[index] if index < len(series_values) else None
        rows.append({"category": category, "values": values})

    return {
        "ref": chart.ref,
        "filename": chart.filename,
        "type": "chart",
        "source_path": chart.source_path,
        "title": _first_text(root, ".//c:title//a:t") or "",
        "series": [str(series["name"]) for series in series_data],
        "categories": categories,
        "data_rows": rows,
    }


def _first_text(root: ET.Element, path: str) -> str:
    values = _texts(root, path)
    return values[0] if values else ""


def _texts(
    root: ET.Element, path: str, namespaces: dict[str, str] | None = None
) -> list[str]:
    return [
        (element.text or "").strip()
        for element in root.findall(path, namespaces or CHART_XML_NS)
        if (element.text or "").strip()
    ]


def _unique_preserving_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    unique = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        unique.append(value)
    return unique


def _number_or_text(value: str) -> object:
    try:
        number = float(value)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def _render_chart_preview(chart: EmbeddedObject, chart_data: dict[str, object]) -> str:
    title = str(chart_data.get("title") or chart.ref)
    lines = [
        f"# Chart Preview {chart.ref}",
        "",
        f"Source part: `{chart.source_path}`",
        f"Title: {title}",
        "",
    ]
    data_rows = chart_data.get("data_rows")
    series = [str(name) for name in chart_data.get("series", [])]
    if isinstance(data_rows, list) and series:
        table_rows = [["Category", *series]]
        for row in data_rows:
            if not isinstance(row, dict):
                continue
            values = row.get("values", {})
            if not isinstance(values, dict):
                values = {}
            table_rows.append([row.get("category", ""), *(values.get(name, "") for name in series)])
        lines.extend(_markdown_table(table_rows))
        lines.append("")
    else:
        lines.extend(["No chart data found.", ""])
    return "\n".join(lines).rstrip() + "\n"


def _render_diagram_preview(diagram: EmbeddedObject, diagram_data: dict[str, object]) -> str:
    lines = [
        f"# SmartArt Preview {diagram.ref}",
        "",
        f"Source part: `{diagram.source_path}`",
        "",
    ]
    texts = [str(text) for text in diagram_data.get("texts", [])]
    if texts:
        lines.extend(f"- {text}" for text in texts)
        lines.append("")
    else:
        lines.extend(["No SmartArt text found.", ""])
    companion_parts = [str(part) for part in diagram_data.get("companion_parts", [])]
    if companion_parts:
        lines.extend(["Companion parts:", ""])
        lines.extend(f"- `{part}`" for part in companion_parts)
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _preview_status(ref: str, json_path: Path | None) -> dict[str, object]:
    return {
        "status": "preview_extracted",
        "preview_path": f"structured/resources/{ref}/preview.md",
        "preview_json_path": (
            f"structured/resources/{ref}/preview.json" if json_path else None
        ),
        "full_parse": False,
    }


def _render_xlsx_preview(embedded: EmbeddedObject, sheets: list[dict[str, Any]]) -> str:
    lines = [
        f"# Resource Preview {embedded.ref}",
        "",
        f"Source file: `{embedded.filename}`",
        "",
    ]
    for sheet in sheets:
        lines.extend(
            [
                f"## Sheet: {sheet['name']}",
                "",
                f"- Rows: {sheet['max_row']}",
                f"- Columns: {sheet['max_column']}",
                "",
            ]
        )
        rows = sheet["preview_rows"]
        if rows:
            lines.extend(_markdown_table(rows))
            lines.append("")
        else:
            lines.extend(["No preview rows found.", ""])
    return "\n".join(lines).rstrip() + "\n"


def _markdown_table(rows: list[list[object | None]]) -> list[str]:
    max_width = max((len(row) for row in rows), default=0)
    if max_width == 0:
        return []
    padded = [row + [None] * (max_width - len(row)) for row in rows]
    rendered = ["| " + " | ".join(_markdown_cell(value) for value in padded[0]) + " |"]
    rendered.append("| " + " | ".join("---" for _ in range(max_width)) + " |")
    for row in padded[1:]:
        rendered.append("| " + " | ".join(_markdown_cell(value) for value in row) + " |")
    return rendered


def _markdown_cell(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|")


def _trim_row(row: list[object | None]) -> list[object | None]:
    while row and row[-1] is None:
        row.pop()
    return row


def _json_safe_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value
