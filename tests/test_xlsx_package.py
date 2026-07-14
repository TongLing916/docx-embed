"""Tests for the enhanced XLSX workbook parser (``parse_xlsx_package``).

Covers:

* Metadata extraction: sheet names, hidden state, merged ranges, hidden rows/columns.
* Cell metadata: formulas, hyperlinks, comments, data types.
* Sub-table splitting: repeated headers, blank-row gaps.
* Markdown rendering: full-row output, intra-cell newline handling.
* Asset annotations in table JSON.
* Framework enrichment placeholders.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment

from edp.xlsx.parser import parse_xlsx_package


def _make_featured_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Financials"
    sheet["A1"] = "Region"
    sheet["B1"] = "Q1"
    sheet["C1"] = "Q2"
    sheet["D1"] = "Total"
    sheet["A2"] = "North"
    sheet["B2"] = 10
    sheet["C2"] = 15
    sheet["D2"] = "=SUM(B2:C2)"
    sheet["A3"] = "South"
    sheet["B3"] = 8
    sheet["C3"] = 11
    sheet["D3"] = "=SUM(B3:C3)"
    sheet.merge_cells("A5:D5")
    sheet["A5"] = "Merged note"
    sheet["A2"].hyperlink = "https://example.com/north"
    sheet["A2"].comment = Comment("Priority market", "EDP")

    hidden = workbook.create_sheet("HiddenData")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "secret"
    sheet.row_dimensions[4].hidden = True
    sheet.column_dimensions["C"].hidden = True

    workbook.save(path)
    return path


def test_parse_xlsx_package_emits_metadata_and_formula_matrix(
    tmp_path: Path,
) -> None:
    workbook_path = _make_featured_workbook(tmp_path / "featured.xlsx")

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "workbook_01")

    metadata = json.loads(
        (package.package_dir / "workbook.json").read_text(encoding="utf-8")
    )
    assert metadata["source_file"] == "featured.xlsx"
    assert [sheet["name"] for sheet in metadata["sheets"]] == [
        "Financials",
        "HiddenData",
    ]
    assert metadata["sheets"][0]["merged_ranges"] == ["A5:D5"]
    assert metadata["sheets"][0]["hidden_rows"] == [4]
    assert metadata["sheets"][0]["hidden_columns"] == ["C"]
    assert metadata["sheets"][1]["state"] == "hidden"

    table_json = json.loads(package.tables[0].json_path.read_text(encoding="utf-8"))
    assert table_json["sheet"] == "Financials"
    assert table_json["cells"]["D2"]["formula"] == "=SUM(B2:C2)"
    assert table_json["cells"]["A2"]["hyperlink"] == "https://example.com/north"
    assert table_json["cells"]["A2"]["comment"] == "Priority market"
    assert table_json["cells"]["B2"]["data_type"] == "n"

    markdown = package.content_path.read_text(encoding="utf-8")
    assert "- Metadata: `workbook.json`" in markdown


def test_intra_cell_newline_does_not_break_markdown_table(
    tmp_path: Path,
) -> None:
    """Embedded newlines render as ``<br>`` so the table stays one row per record."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Wrap"
    sheet["A1"] = "Item"
    sheet["B1"] = "Detail"
    sheet["A2"] = "Test Item"
    sheet["B2"] = "Line one\nLine two\nLine three"

    workbook_path = tmp_path / "wrap.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "wrap_01")

    # Fidelity: raw cell value keeps the literal newlines.
    table_json = json.loads(package.tables[0].json_path.read_text(encoding="utf-8"))
    assert table_json["cells"]["B2"]["value"] == "Line one\nLine two\nLine three"
    assert "\n" in table_json["rows"][1][1]

    # Portability: the Markdown preview uses <br> so the row stays on one line.
    markdown = package.content_path.read_text(encoding="utf-8")
    table_lines = [
        line
        for line in markdown.splitlines()
        if line.startswith("|") and "---" not in line
    ]
    assert len(table_lines) == 2
    assert "Line one<br>Line two<br>Line three" in table_lines[1]
    assert "\n" not in table_lines[1]


def test_split_sheet_into_multiple_tables_by_repeated_headers(
    tmp_path: Path,
) -> None:
    """A sheet with repeated header rows (≥3 occurrences) should produce multiple sub-tables."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inspection"
    sheet.append(["Component", "Check Point", "Standard"])
    sheet.append(["Pantograph", "Surface", "No damage"])
    sheet.append(["Pantograph", "Insulator", "No cracks"])
    sheet.append(["Component", "Check Point", "Standard"])
    sheet.append(["Traction Motor", "Stator", "Good insulation"])
    sheet.append(["Component", "Check Point", "Standard"])
    sheet.append(["Brake System", "Pad Thickness", ">=10mm"])

    workbook_path = tmp_path / "split.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "split_01")

    assert len(package.tables) == 3
    assert package.tables[0].csv_path.name == "table_001.csv"
    assert package.tables[1].csv_path.name == "table_002.csv"
    assert package.tables[2].csv_path.name == "table_003.csv"

    t1 = json.loads(package.tables[0].json_path.read_text(encoding="utf-8"))
    assert t1["sheet"] == "Inspection"
    assert t1["start_row"] == 1
    assert t1["end_row"] == 3
    assert t1["boundary_signals"] == ["header_repeat"]
    assert len(t1["rows"]) == 3

    t2 = json.loads(package.tables[1].json_path.read_text(encoding="utf-8"))
    assert t2["sheet"] == "Inspection"
    assert t2["start_row"] == 4
    assert t2["boundary_signals"] == ["header_repeat"]
    assert len(t2["rows"]) == 2

    t3 = json.loads(package.tables[2].json_path.read_text(encoding="utf-8"))
    assert t3["sheet"] == "Inspection"
    assert t3["start_row"] == 6
    assert t3["boundary_signals"] == []
    assert len(t3["rows"]) == 2


def test_split_sheet_by_blank_row_gap(tmp_path: Path) -> None:
    """Two consecutive blank rows should produce a table boundary."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["A", "B"])
    sheet.append(["1", "2"])
    sheet.append([])
    sheet.append([])
    sheet.append(["C", "D"])
    sheet.append(["3", "4"])

    workbook_path = tmp_path / "blanks.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "blanks_01")

    assert len(package.tables) == 2
    t1 = json.loads(package.tables[0].json_path.read_text(encoding="utf-8"))
    assert t1["boundary_signals"] == ["blank_row_gap"]
    assert t1["end_row"] == 2

    t2 = json.loads(package.tables[1].json_path.read_text(encoding="utf-8"))
    assert t2["start_row"] == 5


def test_single_blank_row_does_not_split(tmp_path: Path) -> None:
    """A single blank row between data rows is not a table boundary."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["A", "B"])
    sheet.append(["1", "2"])
    sheet.append([])
    sheet.append(["3", "4"])

    workbook_path = tmp_path / "single_blank.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "single_01")
    assert len(package.tables) == 1


def test_full_row_rendering_in_content_md(tmp_path: Path) -> None:
    """content.md renders all rows, not just the first 10."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Long"
    sheet.append(["Index", "Value"])
    for i in range(1, 21):
        sheet.append([str(i), str(i * 10)])

    workbook_path = tmp_path / "long.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "long_01")
    markdown = package.content_path.read_text(encoding="utf-8")

    table_lines = [
        line
        for line in markdown.splitlines()
        if line.startswith("|") and "---" not in line
    ]
    assert len(table_lines) == 21
    assert "| 20 |" in markdown


def test_prune_blank_rows_and_columns_removes_noise(tmp_path: Path) -> None:
    """Entirely blank rows and columns within a sub-table are removed."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sparse"
    sheet.append(["Component", None, "Standard", None])
    sheet.append(["Part A", None, "No damage", None])
    sheet.append([None, None, None, None])
    sheet.append(["Part B", None, "Good", None])
    sheet.append(["Component", None, "Standard", None])
    sheet.append(["Part C", None, "OK", None])
    sheet.append(["Component", None, "Standard", None])
    sheet.append(["Part D", None, "Fine", None])

    workbook_path = tmp_path / "sparse.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "sparse_01")
    assert len(package.tables) == 3

    for table in package.tables:
        j = json.loads(table.json_path.read_text(encoding="utf-8"))
        for row in j["rows"]:
            assert len(row) <= 2, f"Row has {len(row)} cells, expected ≤2: {row[:5]}"
            assert None not in row, f"Row contains None after pruning: {row}"

    j1 = json.loads(package.tables[0].json_path.read_text(encoding="utf-8"))
    assert len(j1["rows"]) == 3

    with package.tables[0].csv_path.open(newline="", encoding="utf-8") as f:
        reader = list(csv.reader(f))
    assert len(reader) == 3
    for row in reader:
        assert len(row) == 2, f"CSV row has {len(row)} cols: {row[:5]}"


def test_framework_result_placeholder_in_table_json(tmp_path: Path) -> None:
    """Every table JSON carries a ``framework_result`` placeholder."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "T1"
    sheet.append(["A", "B"])
    sheet.append(["1", "2"])

    workbook_path = tmp_path / "fw.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "fw_01")

    table_json = json.loads(
        package.tables[0].json_path.read_text(encoding="utf-8")
    )
    assert table_json["framework_result"] == {
        "unstructured": None,
        "docling": None,
        "mineru": None,
    }


def test_asset_annotations_in_table_json(tmp_path: Path) -> None:
    """Table JSON surfaces asset count and per-row asset filter."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Assets"
    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    sheet["A2"] = "X"
    sheet["B2"] = 100
    sheet["A3"] = "Y"
    sheet["B3"] = 200

    chart = BarChart()
    chart.add_data(
        Reference(sheet, min_col=2, max_col=2, min_row=1, max_row=3),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D2")

    workbook_path = tmp_path / "assets.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "assets_01")

    assert len(package.tables) == 1
    table_json = json.loads(
        package.tables[0].json_path.read_text(encoding="utf-8")
    )
    assert table_json["chart_count"] == 1
    assert table_json["image_count"] == 0
    assert len(table_json["assets"]["charts"]) == 1


def test_empty_workbook_warns(tmp_path: Path) -> None:
    """A workbook with only empty sheets produces a warning."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Empty"

    workbook_path = tmp_path / "empty.xlsx"
    workbook.save(workbook_path)

    package = parse_xlsx_package(workbook_path, tmp_path / "pkg", "empty_01")
    assert len(package.tables) == 0
    assert any("non-empty sheets" in w for w in package.warnings)


def test_demo_workbook_parses_all_sheets(demo_xlsx_path: Path, tmp_path: Path) -> None:
    """The pre-built demo workbook produces tables for visible sheets + metadata."""

    package = parse_xlsx_package(demo_xlsx_path, tmp_path / "pkg", "demo")

    # At least 3 tables (Sales Summary + Inventory + Data Types, _Config is hidden).
    assert len(package.tables) >= 3

    sheet_names = {t.sheet_name for t in package.tables}
    assert "Sales Summary" in sheet_names
    assert "Inventory" in sheet_names
    assert "Data Types" in sheet_names

    # Metadata
    meta = json.loads((package.package_dir / "workbook.json").read_text(encoding="utf-8"))
    assert meta["sheet_count"] == 4
    assert meta["sheets"][2]["state"] == "hidden"  # _Config sheet

    # Sales summary table has formulas
    sales_table = next(
        t for t in package.tables if t.sheet_name == "Sales Summary"
    )
    sales_json = json.loads(sales_table.json_path.read_text(encoding="utf-8"))
    # Annual Total column (F) should have formula cells
    formula_cells = {
        coord: c
        for coord, c in sales_json.get("cells", {}).items()
        if "formula" in c
    }
    assert len(formula_cells) >= 5  # 5 region rows + 1 total row

    # Chart was extracted
    assert len(sales_json.get("assets", {}).get("charts", [])) >= 1

    # Inventory has a merged note cell
    inv_table = next(
        t for t in package.tables if t.sheet_name == "Inventory"
    )
    inv_json = json.loads(inv_table.json_path.read_text(encoding="utf-8"))
    stock_value_cells = {
        coord: c
        for coord, c in inv_json.get("cells", {}).items()
        if "formula" in c
    }
    assert len(stock_value_cells) >= 8  # Stock Value = D × E for 8 rows

    # Content Markdown is readable
    content = package.content_path.read_text(encoding="utf-8")
    assert "# Embedded Workbook demo" in content
    assert "Sales Summary" in content
    assert "Inventory" in content
