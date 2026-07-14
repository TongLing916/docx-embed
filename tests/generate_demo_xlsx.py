"""Generate a realistic demo XLSX workbook for testing and documentation.

Produces ``demo_workbook.xlsx`` with:
- Multiple sheets (visible + hidden)
- Formulas with both computed and raw values
- Charts (bar chart)
- Merged cells
- Hyperlinks and comments
- Various data types (text, number, date, boolean)
- Hidden rows and columns

Usage::

    python tests/generate_demo_xlsx.py [output_path]
"""

from __future__ import annotations

from pathlib import Path
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, numbers


def build_demo_workbook(output_path: str | Path) -> Path:
    output = Path(output_path)
    wb = Workbook()

    # ── Sheet 1: Sales Summary ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sales Summary"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center")

    headers = ["Region", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue", "Annual Total"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data = [
        ("North America", 125000, 142000, 158000, 172000),
        ("Europe",        98000,  105000, 112000, 128000),
        ("Asia Pacific",  156000, 168000, 175000, 190000),
        ("Latin America", 45000,  52000,  58000,  61000),
        ("Middle East",   38000,  41000,  44000,  47000),
    ]

    for r, (region, q1, q2, q3, q4) in enumerate(data, 2):
        ws1.cell(row=r, column=1, value=region)
        ws1.cell(row=r, column=2, value=q1)
        ws1.cell(row=r, column=3, value=q2)
        ws1.cell(row=r, column=4, value=q3)
        ws1.cell(row=r, column=5, value=q4)
        # Formula column: =SUM(B{r}:E{r})
        ws1.cell(row=r, column=6).value = f"=SUM(B{r}:E{r})"

    # Merge title row above the table
    ws1.insert_rows(1)
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "Annual Sales Report — 2025"
    ws1["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws1["A1"].alignment = Alignment(horizontal="center")

    # Hyperlink and comment
    ws1["A3"].hyperlink = "https://example.com/reports/north-america"
    ws1["A3"].comment = Comment("Top-performing region in 2025", "Report Author")

    # Number formatting
    for r in range(3, 8):
        for c in range(2, 7):
            ws1.cell(row=r, column=c).number_format = '#,##0'

    # Totals row with formulas
    total_row = len(data) + 3
    ws1.cell(row=total_row, column=1, value="TOTAL")
    ws1.cell(row=total_row, column=1).font = Font(bold=True)
    for c in range(2, 7):
        col_letter = chr(64 + c)  # B, C, D, E, F
        ws1.cell(row=total_row, column=c).value = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
        ws1.cell(row=total_row, column=c).number_format = '#,##0'

    # Chart: Q1-Q4 by region
    chart = BarChart()
    chart.title = "Quarterly Revenue by Region"
    chart.y_axis.title = "Revenue (USD)"
    chart.x_axis.title = "Region"
    chart.style = 10
    data_ref = Reference(ws1, min_col=2, max_col=5, min_row=2, max_row=6)
    cats_ref = Reference(ws1, min_col=1, min_row=3, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 22
    chart.height = 14
    ws1.add_chart(chart, "A10")

    # ── Sheet 2: Product Inventory ────────────────────────────────────────
    ws2 = wb.create_sheet("Inventory")
    ws2.append(["SKU", "Product Name", "Category", "Stock", "Unit Price", "Last Restocked"])
    ws2.append(["SKU-001", "Widget Alpha",   "Widgets",  450,  12.99, date(2025, 6, 15)])
    ws2.append(["SKU-002", "Widget Beta",    "Widgets",  320,  15.49, date(2025, 6, 20)])
    ws2.append(["SKU-003", "Gadget Pro",     "Gadgets",  180,  89.00, date(2025, 5, 1)])
    ws2.append(["SKU-004", "Gadget Lite",    "Gadgets",  275,  49.99, date(2025, 6, 10)])
    ws2.append(["SKU-005", "Super Connector","Hardware", 600,   3.75, date(2025, 7, 1)])
    ws2.append(["SKU-006", "Mounting Kit",   "Hardware", 150,  22.50, date(2025, 6, 28)])
    ws2.append(["SKU-007", "Cable Pack",     "Hardware", 900,   8.99, date(2025, 7, 5)])
    ws2.append(["SKU-008", "Sensor Module",  "Gadgets",   85, 125.00, date(2025, 4, 15)])

    # Stock value formula column
    ws2["G1"] = "Stock Value"
    for r in range(2, 10):
        ws2.cell(row=r, column=7).value = f"=D{r}*E{r}"
        ws2.cell(row=r, column=7).number_format = '#,##0.00'

    for col_cells in ws2.iter_cols(min_row=1, max_row=1, min_col=1, max_col=7):
        for cell in col_cells:
            cell.font = header_font
            cell.fill = header_fill

    # Merge a note row
    ws2.merge_cells("A12:G12")
    ws2["A12"] = "Note: Stock values are live-computed from unit price × quantity."
    ws2["A12"].font = Font(italic=True, color="666666")

    # ── Sheet 3: Metadata (hidden) ────────────────────────────────────────
    ws3 = wb.create_sheet("_Config")
    ws3.sheet_state = "hidden"
    ws3["A1"] = "Key"
    ws3["B1"] = "Value"
    ws3["A2"] = "generated_by"
    ws3["B2"] = "demo_workbook_generator.py"
    ws3["A3"] = "generated_at"
    ws3["B3"] = datetime.now().isoformat(sep=" ")
    ws3["A4"] = "currency"
    ws3["B4"] = "USD"
    ws3["A5"] = "fiscal_year"
    ws3["B5"] = 2025

    # Hidden row and column in Inventory
    ws2.row_dimensions[9].hidden = True  # Hide Cable Pack row
    ws2.column_dimensions["E"].hidden = True  # Hide Unit Price column

    # ── Sheet 4: Types Demo ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Data Types")
    ws4.append(["Type", "Example Value", "Notes"])
    ws4.append(["Text", "Hello World", "Plain string"])
    ws4.append(["Integer", 42, "Whole number"])
    ws4.append(["Float", 3.14159, "Decimal number"])
    ws4.append(["Boolean", True, "TRUE/FALSE"])
    ws4.append(["Date", date(2025, 12, 31), "ISO format on export"])
    ws4.append(["DateTime", datetime(2025, 6, 15, 14, 30, 0), "ISO 8601"])
    ws4.append(["Percentage", 0.85, "Formatted as 85%"])
    ws4["B7"].number_format = numbers.FORMAT_PERCENTAGE
    ws4.append(["Currency", 1999.99, "Formatted with $ symbol"])
    ws4["B8"].number_format = '"$"#,##0.00'

    for col_cells in ws4.iter_cols(min_row=1, max_row=1, min_col=1, max_col=3):
        for cell in col_cells:
            cell.font = header_font
            cell.fill = header_fill

    # ── Save ──────────────────────────────────────────────────────────────
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "demo_workbook.xlsx"
    result = build_demo_workbook(path)
    print(f"Demo workbook written to: {result.resolve()}")
