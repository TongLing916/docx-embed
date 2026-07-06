"""Build checklist ground-truth scaffolds from DOCX source.

Reads a DOCX package directly and emits the authoritative structural counts
(embedded objects, images, tables, headings) as a JSON scaffold under
``evaluation/groundtruth/<doc_id>.json``. The ``key_texts`` field is left empty
for a human to fill in 2-5 semantic strings per document.

Existing GT files are never overwritten, so manually added ``key_texts`` survive
re-runs.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from io import BytesIO
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree as ET


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from edp.extractor import (
    _normalize_attachment_payload,
    _ordered_asset_relationships,
    _read_ole_preview_image_rel_ids,
    _read_asset_relationships,
    _read_document_relationship_order,
    _without_ole_preview_images,
)


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
DGM_NS = "http://schemas.openxmlformats.org/drawingml/2006/diagram"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
TBL_TAG = f"{{{W_NS}}}tbl"
TBL_GRID_TAG = f"{{{W_NS}}}tblGrid"
GRID_COL_TAG = f"{{{W_NS}}}gridCol"
TR_TAG = f"{{{W_NS}}}tr"
TC_TAG = f"{{{W_NS}}}tc"
P_TAG = f"{{{W_NS}}}p"
P_PR_TAG = f"{{{W_NS}}}pPr"
P_STYLE_TAG = f"{{{W_NS}}}pStyle"
T_TAG = f"{{{W_NS}}}t"
VAL_ATTR = f"{{{W_NS}}}val"
CHART_SER_TAG = f"{{{C_NS}}}ser"
CHART_TITLE_TAG = f"{{{C_NS}}}title"
CHART_TX_TAG = f"{{{C_NS}}}tx"
CHART_CAT_TAG = f"{{{C_NS}}}cat"
CHART_VALUE_TAG = f"{{{C_NS}}}v"
DRAWING_TEXT_TAG = f"{{{A_NS}}}t"
DRAWING_BLIP_TAG = f"{{{A_NS}}}blip"
DIAGRAM_REL_IDS_TAG = f"{{{DGM_NS}}}relIds"
REL_EMBED_ATTR = f"{{{R_NS}}}embed"
REL_DM_ATTR = f"{{{R_NS}}}dm"

HEADING_LEVEL_RE = re.compile(r"Heading\s*(\d+)", re.IGNORECASE)
CN_HEADING_LEVEL_RE = re.compile(r"标题\s*(\d+)")

GT_ROOT = REPO_ROOT / "evaluation" / "groundtruth"


def build_groundtruth(docx_path: Path) -> dict:
    """Extract the structural ground-truth scaffold from a DOCX file."""

    docx_path = Path(docx_path)
    with zipfile.ZipFile(docx_path) as docx:
        warnings: list[str] = []
        relationships = _read_asset_relationships(docx, warnings)
        relationships = _without_ole_preview_images(
            relationships,
            _read_ole_preview_image_rel_ids(docx),
        )
        document_relationships = _read_document_relationships(docx, warnings)
        relationship_order = _read_document_relationship_order(docx, warnings)
        document_xml = docx.read("word/document.xml")
        style_heading_levels = _read_style_heading_levels(docx)
        charts = _charts(docx, warnings)
        smartarts = _smartarts(docx, warnings)

    embedded_objects = _embedded_objects(docx_path, relationships, relationship_order)
    images = _images(relationships, relationship_order)
    tables, nested_tables, nested_table_assets, headings, checkboxes = _parse_body(
        document_xml,
        style_heading_levels,
        document_relationships,
        smartarts,
    )

    return {
        "doc_id": docx_path.stem,
        "source_file": docx_path.name,
        "expected": {
            "embedded_objects": embedded_objects,
            "images": images,
            "tables": tables,
            "nested_tables": nested_tables,
            "checkboxes": checkboxes,
            "charts": charts,
            "smartarts": smartarts,
            "nested_table_assets": nested_table_assets,
            "headings": headings,
            "key_texts": [],
        },
        "auto_warnings": warnings,
    }


def _embedded_objects(
    docx_path: Path,
    relationships: dict[str, dict[str, str]],
    relationship_order: list[str],
) -> list[dict]:
    objects: list[dict] = []
    with zipfile.ZipFile(docx_path) as docx:
        for rel_id in _ordered_asset_relationships(relationships, relationship_order):
            relationship = relationships[rel_id]
            if relationship["kind"] != "attachment":
                continue
            source_path = relationship["source_path"]
            filename = PurePosixPath(source_path).name
            payload = docx.read(source_path)
            filename, payload, _ = _normalize_attachment_payload(
                filename, payload, unsafe_unwrap=True
            )
            suffix = PurePosixPath(filename).suffix.lower()
            entry = {
                "ref": f"attachment_{len(objects) + 1:02d}",
                "source_path": source_path,
                "type": suffix.removeprefix(".") or "attachment",
                "sheets": None,
            }
            if suffix == ".xlsx":
                entry["type"] = "xlsx"
                entry["sheets"] = _count_xlsx_sheets(payload)
            elif suffix == ".bin":
                entry["type"] = "xlsx"
                entry["sheets"] = _count_xlsx_sheets_from_ole(payload)
            objects.append(entry)
    return objects


def _images(
    relationships: dict[str, dict[str, str]],
    relationship_order: list[str],
) -> list[dict]:
    images: list[dict] = []
    for rel_id in _ordered_asset_relationships(relationships, relationship_order):
        relationship = relationships[rel_id]
        if relationship["kind"] != "image":
            continue
        source_path = relationship["source_path"]
        images.append(
            {
                "ref": f"image_{len(images) + 1:02d}",
                "filename": PurePosixPath(source_path).name,
                "source_path": source_path,
            }
        )
    return images


def _read_document_relationships(
    docx: zipfile.ZipFile, warnings: list[str]
) -> dict[str, dict[str, str]]:
    try:
        rels_xml = docx.read("word/_rels/document.xml.rels")
    except KeyError:
        warnings.append("DOCX relationship file is missing: word/_rels/document.xml.rels")
        return {}
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError as exc:
        warnings.append(f"DOCX relationship file is not valid XML: {exc}")
        return {}
    relationships: dict[str, dict[str, str]] = {}
    for rel in root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if not rel_id or not target or rel.attrib.get("TargetMode") == "External":
            continue
        relationships[rel_id] = {
            "source_path": _word_zip_path(target),
            "relationship_type": rel.attrib.get("Type", ""),
        }
    return relationships


def _word_zip_path(target: str) -> str:
    path = PurePosixPath(target)
    if path.is_absolute():
        path = PurePosixPath(str(path).lstrip("/"))
    if str(path).startswith("word/"):
        return str(path)
    return str(PurePosixPath("word") / path)


def _charts(docx: zipfile.ZipFile, warnings: list[str]) -> list[dict]:
    charts: list[dict] = []
    chart_paths = sorted(
        name
        for name in docx.namelist()
        if name.startswith("word/charts/chart") and name.endswith(".xml")
    )
    for index, source_path in enumerate(chart_paths):
        try:
            root = ET.fromstring(docx.read(source_path))
        except ET.ParseError as exc:
            warnings.append(f"Chart XML is not valid: {source_path}: {exc}")
            continue
        title = _chart_title(root)
        series = _chart_series(root)
        categories = _chart_categories(root)
        texts = _unique_non_empty([title, *series, *categories])
        charts.append(
            {
                "index": index,
                "source_path": source_path,
                "title": title,
                "series": series,
                "categories": categories,
                "texts": texts,
            }
        )
    return charts


def _smartarts(docx: zipfile.ZipFile, warnings: list[str]) -> list[dict]:
    smartarts: list[dict] = []
    diagram_paths = sorted(
        name
        for name in docx.namelist()
        if name.startswith("word/diagrams/data") and name.endswith(".xml")
    )
    seen_texts: set[tuple[str, ...]] = set()
    for source_path in diagram_paths:
        try:
            root = ET.fromstring(docx.read(source_path))
        except ET.ParseError as exc:
            warnings.append(f"SmartArt XML is not valid: {source_path}: {exc}")
            continue
        texts = _drawing_texts(root)
        text_key = tuple(texts)
        if texts and text_key not in seen_texts:
            seen_texts.add(text_key)
            smartarts.append(
                {
                    "index": len(smartarts),
                    "source_path": source_path,
                    "texts": texts,
                }
            )
    return smartarts


def _drawing_texts(root: ET.Element) -> list[str]:
    return _unique_non_empty(
        [
            text
            for node in root.iter(DRAWING_TEXT_TAG)
            if (text := (node.text or "").strip())
        ]
    )


def _chart_title(root: ET.Element) -> str:
    title = root.find(f".//{CHART_TITLE_TAG}")
    if title is None:
        return ""
    return " ".join(
        text
        for node in title.iter(DRAWING_TEXT_TAG)
        if (text := (node.text or "").strip())
    )


def _chart_series(root: ET.Element) -> list[str]:
    series: list[str] = []
    for ser in root.iter(CHART_SER_TAG):
        tx = ser.find(CHART_TX_TAG)
        if tx is None:
            continue
        values = [
            text
            for node in tx.iter(CHART_VALUE_TAG)
            if (text := (node.text or "").strip())
        ]
        if values:
            series.append(values[0])
    return _unique_non_empty(series)


def _chart_categories(root: ET.Element) -> list[str]:
    categories: list[str] = []
    for cat in root.iter(CHART_CAT_TAG):
        categories.extend(
            text
            for node in cat.iter(CHART_VALUE_TAG)
            if (text := (node.text or "").strip())
        )
    return _unique_non_empty(categories)


def _unique_non_empty(values: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            unique.append(value)
    return unique


def _parse_body(
    document_xml: bytes,
    style_heading_levels: dict[str, int] | None = None,
    relationships: dict[str, dict[str, str]] | None = None,
    smartarts: list[dict] | None = None,
) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict]]:
    root = ET.fromstring(document_xml)
    tables, nested_tables, nested_table_assets = _parse_tables(
        root,
        relationships or {},
        smartarts or [],
    )
    headings = _parse_headings(root, style_heading_levels or {})
    checkboxes = _parse_checkboxes(root)
    return tables, nested_tables, nested_table_assets, headings, checkboxes


STYLE_TAG = f"{{{W_NS}}}style"
STYLE_ID_ATTR = f"{{{W_NS}}}styleId"
STYLE_NAME_TAG = f"{{{W_NS}}}name"
STYLE_NAME_ATTR = f"{{{W_NS}}}val"
STYLE_BASED_ON_TAG = f"{{{W_NS}}}basedOn"


def _read_style_heading_levels(docx: zipfile.ZipFile) -> dict[str, int]:
    """Map paragraph styleId -> heading level (1-9) by resolving styles.xml.

    A style is a heading when its ``w:name`` is ``heading N`` (English) or
    ``标题 N`` (Chinese). Numeric styleIds are common aliases (e.g. ``1`` ->
    ``heading 1``) and must be resolved through the style table, since styleIds
    like ``10``/``20`` often name ``目录 N`` (TOC) and are not headings.
    """

    try:
        styles_xml = docx.read("word/styles.xml")
    except KeyError:
        return {}
    root = ET.fromstring(styles_xml)

    raw: dict[str, tuple[int | None, str | None]] = {}
    for style in root.findall(STYLE_TAG):
        style_id = style.attrib.get(STYLE_ID_ATTR, "")
        if not style_id:
            continue
        name_el = style.find(STYLE_NAME_TAG)
        name = name_el.attrib.get(STYLE_NAME_ATTR, "") if name_el is not None else ""
        based_el = style.find(STYLE_BASED_ON_TAG)
        based_on = based_el.attrib.get(STYLE_NAME_ATTR) if based_el is not None else None
        raw[style_id] = (_name_heading_level(name), based_on)

    def resolve(style_id: str, seen: set[str] | None = None) -> int | None:
        seen = seen or set()
        if style_id in seen or style_id not in raw:
            return None
        seen.add(style_id)
        level, based_on = raw[style_id]
        if level is not None:
            return level
        return resolve(based_on, seen) if based_on else None

    return {sid: level for sid in raw if (level := resolve(sid)) is not None}


def _name_heading_level(name: str) -> int | None:
    match = HEADING_LEVEL_RE.search(name)
    if match:
        return int(match.group(1))
    match = CN_HEADING_LEVEL_RE.search(name)
    if match:
        return int(match.group(1))
    return None


def _parse_tables(
    root: ET.Element,
    relationships: dict[str, dict[str, str]],
    smartarts: list[dict],
) -> tuple[list[dict], list[dict], list[dict]]:
    tables: list[dict] = []
    nested_tables: list[dict] = []
    nested_table_assets: list[dict] = []
    parents = _parent_map(root)
    smartart_texts_by_path = {
        str(smartart.get("source_path", "")): list(smartart.get("texts", []))
        for smartart in smartarts
    }
    for index, tbl in enumerate(root.iter(TBL_TAG)):
        rows = len(tbl.findall(TR_TAG))
        cols = _table_column_count(tbl)
        depth = _table_depth(tbl, parents)
        entry = {"index": index, "rows": rows, "cols": cols}
        if depth:
            nested_tables.append(
                {
                    **entry,
                    "depth": depth,
                    "texts": _table_texts(tbl),
                }
            )
            nested_table_assets.extend(
                _nested_table_assets(
                    tbl,
                    index,
                    relationships,
                    smartart_texts_by_path,
                )
            )
        else:
            tables.append(entry)
    return tables, nested_tables, nested_table_assets


def _nested_table_assets(
    tbl: ET.Element,
    table_index: int,
    relationships: dict[str, dict[str, str]],
    smartart_texts_by_path: dict[str, list[str]],
) -> list[dict]:
    assets: list[dict] = []
    for blip in tbl.iter(DRAWING_BLIP_TAG):
        rel_id = blip.attrib.get(REL_EMBED_ATTR)
        if not rel_id:
            continue
        assets.append(
            {
                "table_index": table_index,
                "kind": "image",
                "rel_id": rel_id,
                "texts": [],
            }
        )
    for rel_ids in tbl.iter(DIAGRAM_REL_IDS_TAG):
        rel_id = rel_ids.attrib.get(REL_DM_ATTR)
        if not rel_id:
            continue
        source_path = relationships.get(rel_id, {}).get("source_path", "")
        assets.append(
            {
                "table_index": table_index,
                "kind": "smartart",
                "rel_id": rel_id,
                "texts": smartart_texts_by_path.get(source_path, []),
            }
        )
    return assets


def _table_column_count(tbl: ET.Element) -> int:
    grid = tbl.find(TBL_GRID_TAG)
    if grid is not None:
        grid_cols = len(grid.findall(GRID_COL_TAG))
        if grid_cols:
            return grid_cols
    max_cols = 0
    for tr in tbl.findall(TR_TAG):
        max_cols = max(max_cols, len(tr.findall(TC_TAG)))
    return max_cols


def _parent_map(root: ET.Element) -> dict[int, ET.Element]:
    return {id(child): parent for parent in root.iter() for child in parent}


def _table_depth(tbl: ET.Element, parents: dict[int, ET.Element]) -> int:
    depth = 0
    current = parents.get(id(tbl))
    while current is not None:
        if current.tag == TBL_TAG:
            depth += 1
        current = parents.get(id(current))
    return depth


def _table_texts(tbl: ET.Element) -> list[str]:
    texts: list[str] = []
    for text_node in tbl.iter(T_TAG):
        text = (text_node.text or "").strip()
        if text:
            texts.append(text)
    return texts


def _parse_checkboxes(root: ET.Element) -> list[dict]:
    checkboxes: list[dict] = []
    for para in root.iter(P_TAG):
        text = "".join(t.text or "" for t in para.iter(T_TAG)).strip()
        if text and _has_checkbox_marker(text):
            checkboxes.append({"text": text})
    return checkboxes


def _has_checkbox_marker(text: str) -> bool:
    return any(marker in text for marker in ("□", "■", "☑", "☒", "☐"))


def _parse_headings(root: ET.Element, style_heading_levels: dict[str, int]) -> list[dict]:
    headings: list[dict] = []
    for para in root.iter(P_TAG):
        level = _heading_level(para, style_heading_levels)
        if level is None:
            continue
        text = "".join(t.text or "" for t in para.iter(T_TAG)).strip()
        if not text:
            continue
        headings.append({"level": level, "text": text})
    return headings


def _heading_level(para: ET.Element, style_heading_levels: dict[str, int]) -> int | None:
    """Return 1-9 for true heading paragraphs, else None.

    Resolves the paragraph's ``pStyle`` through ``style_heading_levels`` (built
    from ``word/styles.xml``), so numeric styleIds that alias ``heading N`` are
    recognized while ``目录 N`` (TOC) styles are rejected.
    """

    p_pr = para.find(P_PR_TAG)
    if p_pr is None:
        return None
    p_style = p_pr.find(P_STYLE_TAG)
    if p_style is None:
        return None
    style_id = p_style.attrib.get(VAL_ATTR, "")
    if not style_id:
        return None
    if style_id in style_heading_levels:
        return style_heading_levels[style_id]
    return None


def _count_xlsx_sheets(payload: bytes) -> int | None:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        try:
            return len(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception:
        return None


def _count_xlsx_sheets_from_ole(payload: bytes) -> int | None:
    from edp.extractor import _extract_xlsx_from_ole_payload

    xlsx_payload = _extract_xlsx_from_ole_payload(payload)
    if xlsx_payload is None:
        return None
    return _count_xlsx_sheets(xlsx_payload)


def write_groundtruth(docx_path: Path, gt_root: Path = GT_ROOT, force: bool = False) -> tuple[Path, bool]:
    gt_root.mkdir(parents=True, exist_ok=True)
    out_path = gt_root / f"{Path(docx_path).stem}.json"
    if out_path.exists() and not force:
        return out_path, False
    gt = build_groundtruth(docx_path)
    out_path.write_text(json.dumps(gt, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path, True


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build checklist ground-truth scaffolds from DOCX files.")
    parser.add_argument("inputs", nargs="+", type=Path, help="DOCX file(s) or directory(ies) to scan.")
    parser.add_argument("--gt-root", type=Path, default=GT_ROOT, help="Output directory for GT JSON files.")
    parser.add_argument("--force", action="store_true", help="Overwrite existing GT files (drops manual key_texts).")
    args = parser.parse_args(argv)

    docx_files: list[Path] = []
    for raw in args.inputs:
        if raw.is_dir():
            docx_files.extend(sorted(raw.glob("*.docx")))
        elif raw.suffix.lower() == ".docx":
            docx_files.append(raw)

    if not docx_files:
        print("No DOCX files found.", file=sys.stderr)
        return 1

    for docx in docx_files:
        try:
            out, wrote = write_groundtruth(docx, args.gt_root, force=args.force)
        except (zipfile.BadZipFile, ET.ParseError) as exc:
            print(f"SKIP {docx.name}: {exc}", file=sys.stderr)
            continue
        rel = out.relative_to(REPO_ROOT) if out.is_relative_to(REPO_ROOT) else out
        print(f"{'wrote' if wrote else 'skip'} {rel}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
